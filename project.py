# project.py — CS50P Final Project
# VOI Volume Aligner: aligns user VOI volumes (imported from MIPAV) into the
# correct order based on Baseline reference cell positions, for every
# sheet/subject in an Excel workbook.

from openpyxl import load_workbook

# excel structure layout:
# ============================================================
BASELINE_COL = 2                # column B -> Baseline
TITLE_COL = 4                   # column D -> YR1/YR2 (user) VOI Titles "e.x 18_adjusYR1_PR"
DATA_COL = 5                    # column E -> VOI volume
HEADER_ROW = 1                  # row headers are on (Baseline, Year, Data titles); data starts the row after
SORTED_VOI_NAME = 7             # column header for final user VOI names that are in correct order w/ baseline
SORTED_VOI_VOLUME = 8           # column header containing respective volumes
# ============================================================


def extract_leading_number(value):
    """Extracts only the leading number of a VOI name (e.g. '8_PR' -> 8)."""
    if isinstance(value, (int, float)):
        number = int(value)
    else:
        extracted_number = ""
        for character in value:
            if character != "_":
                extracted_number += character
            else:
                break
        number = int(extracted_number)
    return number


def parse_baseline_column(sheet, first_row, last_row):
    """Reads the Baseline column (B) and returns a list of extracted numbers,
    in sheet order, skipping blank rows."""
    baseline = []
    for row in sheet.iter_rows(first_row, last_row):
        VOI_name = row[BASELINE_COL - 1].value
        if VOI_name is None:
            continue
        baseline.append(extract_leading_number(VOI_name))
    return baseline


def parse_user_column(sheet, first_row, last_row):
    """Reads the user VOI titles/volumes (columns D, E) and returns a dict
    keyed by extracted number: {number: [original_title, volume]}."""
    user_extracted = {}
    for row in sheet.iter_rows(first_row, last_row):
        VOI_name = row[TITLE_COL - 1].value
        if VOI_name is None:
            continue
        VOI_value = row[DATA_COL - 1].value
        VOI_name_extracted = extract_leading_number(VOI_name)
        user_extracted[VOI_name_extracted] = [VOI_name, VOI_value]
    return user_extracted


def align_voi_data(baseline, user_extracted):
    """Matches each baseline number to the user's VOI title/volume.
    Returns (aligned_titles, aligned_values, used_numbers). Unmatched
    baseline positions are filled with 'N/V'."""
    aligned_titles = []
    aligned_values = []
    used_numbers = set()

    for base_number in baseline:
        if base_number in user_extracted:
            title, value = user_extracted[base_number]
            aligned_titles.append(title)
            aligned_values.append(value)
            used_numbers.add(base_number)
        else:
            aligned_titles.append("N/V")
            aligned_values.append("N/V")

    return aligned_titles, aligned_values, used_numbers


def find_new_entries(user_extracted, used_numbers):
    """Finds user VOIs with no matching baseline number.
    Returns (new_titles, new_values)."""
    new_titles = []
    new_values = []
    for number in user_extracted:
        if number not in used_numbers:
            title, value = user_extracted[number]
            new_titles.append(title)
            new_values.append(value)
    return new_titles, new_values


def write_results(sheet, first_row, last_row, aligned_titles, aligned_values, new_titles, new_values):
    """Writes the aligned VOI data (columns G, H) and appends any new,
    unmatched VOI entries after the last existing row."""
    for i in range(len(aligned_titles)):
        row_number = first_row + i
        sheet.cell(row=row_number, column=SORTED_VOI_NAME).value = aligned_titles[i]
        sheet.cell(row=row_number, column=SORTED_VOI_VOLUME).value = aligned_values[i]

    for i in range(len(new_titles)):
        row_number = last_row + 1 + i
        sheet.cell(row=row_number, column=SORTED_VOI_NAME).value = new_titles[i]
        sheet.cell(row=row_number, column=SORTED_VOI_VOLUME).value = new_values[i]


def main(input_path, output_path):
    """Loops through every sheet (subject) in the workbook, aligns each
    subject's VOI data to their Baseline, and saves the result."""
    VOI_excel = load_workbook(input_path)

    for sheet_name in VOI_excel.sheetnames:
        sheet = VOI_excel[sheet_name]
        first_row = HEADER_ROW + 1
        last_row = sheet.max_row

        baseline = parse_baseline_column(sheet, first_row, last_row)
        user_extracted = parse_user_column(sheet, first_row, last_row)
        aligned_titles, aligned_values, used_numbers = align_voi_data(baseline, user_extracted)
        new_titles, new_values = find_new_entries(user_extracted, used_numbers)

        write_results(sheet, first_row, last_row, aligned_titles, aligned_values, new_titles, new_values)

    VOI_excel.save(output_path)


if __name__ == "__main__":
    input_path = "CS50p_VOI_template.xlsx"
    output_path = "CS50p_VOI_output.xlsx"
    main(input_path, output_path)